# -*- coding: utf-8 -*-
"""_data/projects.yml -> 엑셀 파일로 정리해서 내보낸다.

    python scripts/export_projects_xlsx.py                 # 바탕화면에 저장
    python scripts/export_projects_xlsx.py "C:/경로/파일.xlsx"

엑셀은 저장소에 커밋하지 않는다 (.gitignore 의 *.xlsx).
데이터의 원본은 _data/projects.yml 이고, 엑셀은 보기용 사본이다.
"""
import io
import os
import sys
import datetime
import yaml

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl 이 필요하다:  python -m pip install openpyxl")
    sys.exit(1)

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

STATUS_KO = {"ongoing": "진행중", "terminated": "종료"}

HEADERS = [
    ("no", "번호", 6),
    ("status", "상태", 9),
    ("title_ko", "과제명 (국문)", 52),
    ("title_en", "과제명 (영문)", 58),
    ("role", "세부과제 / 역할", 30),
    ("funder", "지원기관", 22),
    ("ministry", "소관 부처", 20),
    ("start", "시작일", 13),
    ("end", "종료일", 13),
    ("months", "기간(개월)", 11),
]


def months_between(a, b):
    if not a or not b:
        return None
    try:
        d1 = datetime.datetime.strptime(a, "%Y.%m.%d")
        d2 = datetime.datetime.strptime(b, "%Y.%m.%d")
    except ValueError:
        return None
    return (d2.year - d1.year) * 12 + (d2.month - d1.month) + (1 if d2.day >= d1.day else 0)


def main():
    dest = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.expanduser("~"), "Desktop", "ECML_projects.xlsx")

    projects = yaml.safe_load(
        io.open(os.path.join(ROOT, "_data", "projects.yml"), encoding="utf-8").read()) or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    head_fill = PatternFill("solid", fgColor="062B45")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D5E6ED")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ongoing_fill = PatternFill("solid", fgColor="EEF9FC")

    for i, (_, label, width) in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 26

    for r, p in enumerate(projects, start=2):
        row = {
            "no": r - 1,
            "status": STATUS_KO.get(p.get("status"), p.get("status")),
            "title_ko": p.get("title_ko"),
            "title_en": p.get("title_en"),
            "role": p.get("role"),
            "funder": p.get("funder"),
            "ministry": p.get("ministry"),
            "start": p.get("start"),
            "end": p.get("end"),
            "months": months_between(p.get("start"), p.get("end")),
        }
        for i, (key, _, _) in enumerate(HEADERS, start=1):
            c = ws.cell(row=r, column=i, value=row[key])
            c.border = border
            c.alignment = Alignment(
                vertical="top",
                horizontal="center" if key in ("no", "status", "start", "end", "months") else "left",
                wrap_text=key in ("title_ko", "title_en", "role"))
            c.font = Font(size=10, bold=(key == "title_ko"))
            if p.get("status") == "ongoing":
                c.fill = ongoing_fill
        ws.row_dimensions[r].height = 46

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(HEADERS)), len(projects) + 1)

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    from collections import Counter
    rows = [("총 과제 수", len(projects)),
            ("진행중", sum(1 for p in projects if p.get("status") == "ongoing")),
            ("종료", sum(1 for p in projects if p.get("status") == "terminated")),
            ("", ""),
            ("지원기관별", "")]
    for k, v in Counter(p.get("funder") for p in projects).most_common():
        rows.append(("  " + str(k), v))
    rows += [("", ""), ("생성 시각", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
             ("원본", "_data/projects.yml")]
    for r, (a, b) in enumerate(rows, start=1):
        ws2.cell(row=r, column=1, value=a).font = Font(size=10, bold=(r == 1 or a in ("지원기관별",)))
        ws2.cell(row=r, column=2, value=b).font = Font(size=10)
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 14

    wb.save(dest)
    print("저장: %s" % dest)
    print("  과제 %d건 (진행중 %d / 종료 %d)"
          % (len(projects),
             sum(1 for p in projects if p.get("status") == "ongoing"),
             sum(1 for p in projects if p.get("status") == "terminated")))
    return 0


if __name__ == "__main__":
    sys.exit(main())
